"""Extract energy data from LAMPIRAN 1 Excel submission."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl


def _val(row: tuple, idx: int, default=None):
    try:
        v = row[idx]
        return v if v is not None else default
    except IndexError:
        return default


def _fmt_month(val) -> str:
    if isinstance(val, datetime):
        return val.strftime("%b %Y")
    return str(val).strip() if val else ""


def _read_monthly(ws, data_row: int, header_row: int, start_col: int) -> tuple[list, list, float]:
    """Read month labels from header_row and values from data_row starting at start_col."""
    headers = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    data    = list(ws.iter_rows(min_row=data_row,   max_row=data_row,   values_only=True))[0]

    months: list[str]   = []
    values: list[float] = []
    total = 0.0

    for i in range(start_col, len(headers)):
        h = headers[i]
        if h is None:
            continue
        lbl = _fmt_month(h)
        if "total" in str(lbl).lower() or "total" in str(h).lower():
            v = data[i] if i < len(data) else None
            if isinstance(v, (int, float)):
                total = float(v)
            break
        if lbl:
            months.append(lbl)
            v = data[i] if i < len(data) else None
            values.append(float(v) if isinstance(v, (int, float)) else 0.0)

    if not total and values:
        total = sum(values)

    return months, values, total


def extract_bei_excel(xlsx_path: Path) -> dict[str, Any]:
    """Extract LAMPIRAN 1(A/B/C) energy data from the Excel submission."""
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    result: dict[str, Any] = {}

    # ── LAMPIRAN 1(A): Supply authority ───────────────────────────────────────
    for name in ["LAMPIRAN 1 (A)", "LAMPIRAN 1(A)"]:
        if name in wb.sheetnames:
            ws   = wb[name]
            row6 = list(ws.iter_rows(min_row=6, max_row=6, values_only=True))[0]
            months, monthly_a, total_a = _read_monthly(ws, 6, 5, 10)
            result.update({
                "building_name": str(_val(row6, 6, "") or "").strip(),
                "address": " ".join(filter(None, [
                    str(_val(row6, 7, "") or ""),
                    str(_val(row6, 8, "") or ""),
                    str(_val(row6, 9, "") or ""),
                ])).strip(),
                "tnb_account":  str(_val(row6, 2, "") or "").strip(),
                "supply_auth":  str(_val(row6, 1, "") or "").strip(),
                "months":       months,
                "monthly_a":    monthly_a,
                "total_a":      total_a,
                "period_label": f"{months[0]} to {months[-1]}" if months else "",
            })
            break

    # ── LAMPIRAN 1(B): Tenant consumption ─────────────────────────────────────
    for name in ["LAMPIRAN 1 (B)", "LAMPIRAN 1(B)"]:
        if name in wb.sheetnames:
            ws = wb[name]
            _, monthly_b, total_b = _read_monthly(ws, 6, 5, 17)
            result.update({"monthly_b": monthly_b, "total_b": total_b})
            break

    # ── LAMPIRAN 1(C): Net own consumption ────────────────────────────────────
    for name in ["LAMPIRAN 1 (C)", "LAMPIRAN 1(C)"]:
        if name in wb.sheetnames:
            ws   = wb[name]
            row6 = list(ws.iter_rows(min_row=6, max_row=6, values_only=True))[0]
            if not result.get("building_name"):
                result["building_name"] = str(_val(row6, 1, "") or "").strip()
            _, monthly_c, total_c = _read_monthly(ws, 6, 5, 5)
            result.update({"monthly_c": monthly_c, "total_c": total_c})
            break

    # Fallback: compute C = A - B if sheet was blank
    if not result.get("total_c") and result.get("total_a"):
        result["total_c"]   = result.get("total_a", 0) - result.get("total_b", 0)
        result["monthly_c"] = [
            a - b for a, b in zip(result.get("monthly_a", []), result.get("monthly_b", []))
        ]

    result.setdefault("total_b", 0)
    result.setdefault("monthly_b", [])
    result.setdefault("total_c", result.get("total_a", 0))
    result.setdefault("monthly_c", result.get("monthly_a", []))

    # Trim to latest 12 months
    months = result.get("months", [])
    n = len(months)
    if n > 12:
        months    = months[-12:]
        monthly_a = result.get("monthly_a", [])
        monthly_b = result.get("monthly_b", [])
        monthly_c = result.get("monthly_c", [])
        if len(monthly_a) == n: monthly_a = monthly_a[-12:]
        if len(monthly_b) == n: monthly_b = monthly_b[-12:]
        if len(monthly_c) == n: monthly_c = monthly_c[-12:]
        result.update({
            "months":       months,
            "monthly_a":    monthly_a,
            "monthly_b":    monthly_b,
            "monthly_c":    monthly_c,
            "total_a":      sum(monthly_a),
            "total_b":      sum(monthly_b),
            "total_c":      sum(monthly_c),
            "period_label": f"{months[0]} to {months[-1]}" if months else "",
        })

    return result
